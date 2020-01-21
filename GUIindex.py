"""
- 10/01/2020 - This program is still in its draft mode, and needs some SERIOUS redesign and a better
    organization of its contents. Keep that in mind, and try to register new alterations here for 
    future reference (at least date and a Code pinting to a readme file with more descriptive info)
andiwillsaythatweshouldtakeadaytobreakawayfromallthepainourbrainhasmadethegameisnotplayedalone 
andiwillsaythatweshouldtakeamomentandholditandkeepitfrozenandknowthatlifehasahopefulundertone
"""
from tkinter import filedialog
from tkinter import *
from tkinter import ttk
from openpyxl.styles import *
from openpyxl.worksheet.page import *
import openpyxl as ol
import datetime as dt
import pandas as pd 
import numpy as np 



df = {}

class Application:
    def __init__(self, myapp):
        self.myapp = myapp
        myapp.title('Bom dia')

        self.label = Label(myapp, text='Relatórios')
        self.label.pack()

        self.entry_path = Entry(myapp, text='Entre com o caminho desejado: ')
        self.entry_path.pack()
        self.button = Button(myapp, text='Encontrar Arquivo', command=self.main_fun)
        self.button.pack()

        """
        Melhor criar um evento que quando disparado, ai possibilita essa opção, junto com outras
        """
        #self.button_comp_sea = Button(myapp, text='Compared Search', command=self.main_comp)
        #self.button_comp_sea.pack()

        myapp.geometry('520x360')

        
    def main_fun(self):
        
        self.myapp.filename = filedialog.askopenfilename(initialdir="Desktop", title='Select File', filetypes = (('excel files', '*.xls'), ('excel 2010', '*.xlsx'), ('all files', '*.*')))
        
        self.lbl_confirm = Label(self.myapp, text = ('caminho em uso: ', (self.myapp.filename)))
        self.lbl_confirm.pack()

        self.lbl_confirm01 = Label(self.myapp, text = 'path found, File Ok')
        self.lbl_confirm01.pack()

        self.button_100 = Button(self.myapp, text='Cem Mais Vendidos ', command=self.toexcel_100)
        self.button_100.pack()

        self.button_ab30 = Button(self.myapp, text='Abaixo de 30% ', command=self.toexcel_30)
        self.button_ab30.pack()

        self.button_30_comp = Button(self.myapp, text='30% Comparado', command=self.toexcel_30comp)
        self.button_30_comp.pack()

    def toexcel_100(self):
        path_to_use = self.myapp.filename
        
        df = pd.read_excel(path_to_use, 
            usecols=[2, 3, 5, 8, 16], 
            names=['cod', 'prod', 'qtd', 'cst_un', 'vnd_un'],
            skiprows=7)
        
        cod = df['cod']
        prod = df['prod']
        qtd = df['qtd']
        cst = df['cst_un']
        vnd = df['vnd_un']

        df_dict = {'cod' : cod, 
            'prod': prod,
            'qtd' : qtd,
            'cst_un': cst,
            'vnd' : vnd,
            'cst_to': cst * qtd,
            'vnd_to': vnd * qtd,
            'margem' :  round(((vnd - cst)/vnd) * 100, 2)}


        df_100 = pd.DataFrame(df_dict)
        
        df_100['vnd_to'] = df_100['vnd_to'].astype(float)
        cem_mais = df_100.sort_values(by='vnd_to', ascending = False)[0:101]
        
        output_path = filedialog.asksaveasfilename(initialdir="Desktop", title='Select Path to Save', filetypes = (('excel 2010', '*.xlsx'), ('all files', '*.*')), defaultextension='.xlsx')
        cem_mais.to_excel(str(output_path), header= ['Código', 'Produto', 'QTD', 'Custo', 'Venda', 'Custo Total', 'Venda Total', 'Margem'], startrow=2 )

    
        wb = ol.load_workbook(output_path)
        ws = wb.active
        
        


        relatorio = NamedStyle(name='relatorio')
        relatorio.font = Font(bold=True, name='Tahoma', size=16)
        bd = Side(style='thin', color="000000")
        relatorio.border = Border(left=bd, right=bd, top=bd, bottom=bd)

        g105 = ws['G105']
        h105 = ws['H105']
        i105 = ws['I105']
        g106 = ws['G106']
        h106 = ws['H106']
        i106 = ws['I106']

        ws['G105'] = 'Total de Custo'
        ws['H105'] = 'Total de Venda'
        ws['I105'] = 'Margem'
        ws['G106'] = "=SUM(G4:G104)"
        ws['H106'] = "=SUM(H4:H104)"
        ws['I106'] = "=((H106-G106)/G106) * 100"

        for cell in ws['B']:
            cell.style = relatorio
        for cell in ws['C']:
            cell.style = relatorio
        for cell in ws['D']:
            cell.style = relatorio
        for cell in ws['E']:
            cell.style = relatorio
            cell.number_format = '#,##0.00'
        for cell in ws['F']:
            cell.style = relatorio
            cell.number_format = '#,##0.00'
        for cell in ws['G']:
            cell.style = relatorio
            cell.number_format = '#,##0.00'
        for cell in ws['H']:
            cell.style = relatorio
            cell.number_format = '#,##0.00'
        for cell in ws['I']:
            cell.style = relatorio
            

        ws.merge_cells('A1:I1')
        a1 = ws['A1']
        a1.font = Font(name='Tahoma', size=26, bold=True)
        c1 = ws['C1']
        ws.column_dimensions['B'].width = 13
        ws.column_dimensions['C'].width = 70
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 13
        ws.column_dimensions['G'].width = 21
        ws.column_dimensions['H'].width = 21
        ws.column_dimensions['I'].width = 23
  
        ws.merge_cells('A2:B2')
        ws.merge_cells('C2:I2')

        
        a2 = ws['A2']
        ws['A2'] = 'A/C: Diretor'
        a2.font = Font(name='Arial', size=15, bold=True)
        ws['C2'] = dt.datetime.today()
        c2=ws['C2']
        c2.font = Font(name='Arial', size=15, bold=True)
        ws['A1'] = '→OS 100 MAIS VENDIDOS  -  SOBERANO FILIAL 06'
        c2.number_format = 'dd-mm-yyyy'
        
        
        
        ws.border = Border(left=Side(border_style='thin'))
        ws.print_area = 'A1:I104'
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToHeight = 2
        ws.page_setup.fitToWidth = True
        wb.save(output_path)

    def toexcel_30(self):
        path_to_use30 = self.myapp.filename
        df_1 = pd.read_excel(path_to_use30, 
            usecols=[2, 3, 5, 8, 16], 
            names=['cod', 'prod', 'qtd', 'cst_un', 'vnd_un'],
            skiprows=7)
        
        cod = df_1['cod']
        prod = df_1['prod']
        qtd = df_1['qtd']
        cst = df_1['cst_un']
        vnd = df_1['vnd_un']

        df_dict = {'cod' : cod, 
            'prod': prod,
            'qtd' : qtd,
            'cst_un': cst,
            'vnd' : vnd,
            'cst_to': cst * qtd,
            'vnd_to': vnd * qtd,
            'margem' : round(((vnd - cst)/vnd) * 100, 2)}

        df_30 = pd.DataFrame(df_dict)
        
        df_30['margem'] = df_30['margem'].astype(float)
        abaixo30 = df_30.sort_values(by='margem', ascending = True)
        abaixos = abaixo30[abaixo30['margem'] < 30] 
        output_path30 = filedialog.asksaveasfilename(initialdir="Desktop", title='Select Path to Save', filetypes = (('excel 2010', '*.xlsx'), ('all files', '*.*')), defaultextension='.xlsx')
        abaixos.to_excel(str(output_path30), header = ['Código', 'Produto', 'QTD', 'Custo', 'Venda', 'Custo Total', 'Venda Total', 'Margem'], startrow=2 )
        

        

        #styling
        wb = ol.load_workbook(output_path30)
        ws = wb.active
        
        


        relatorio = NamedStyle(name='relatorio')
        relatorio.font = Font(bold=True, name='Tahoma', size=16)
        bd = Side(style='thin', color="000000")
        relatorio.border = Border(left=bd, right=bd, top=bd, bottom=bd)
        top_title = ws['A1']
        top_title.fill
        



        for cell in ws['B']:
            cell.style = relatorio
        for cell in ws['C']:
            cell.style = relatorio
        for cell in ws['D']:
            cell.style = relatorio
        for cell in ws['E']:
            cell.style = relatorio
            cell.number_format = '#,##0.00'
        for cell in ws['F']:
            cell.style = relatorio
            cell.number_format = '#,##0.00'
        for cell in ws['G']:
            cell.style = relatorio
            cell.number_format = '#,##0.00'
        for cell in ws['H']:
            cell.style = relatorio
            cell.number_format = '#,##0.00'
        for cell in ws['I']:
            cell.style = relatorio
        

        ws.merge_cells('A1:I1')
        a1 = ws['A1']
        a1.font = Font(name='Arial', size=26, bold=True)
        c1 = ws['C1']
        ws.column_dimensions['B'].width = 13
        ws.column_dimensions['C'].width = 70
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 13
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 17
        ws.column_dimensions['I'].width = 23

        ws.merge_cells('A2:B2')
        ws.merge_cells('C2:I2')

        a2 = ws['A2']
        ws['A2'] = 'A/C: Diretor'
        a2.font = Font(name='Arial', size=15, bold=True)
        ws['C2'] = dt.datetime.today()
        c2=ws['C2']
        c2.font = Font(name='Arial', size=15, bold=True)
        c2.number_format = 'dd-mm-yyyy'
        ws['A1'] = '→ABAIXO DE 30%  -  SOBERANO FILIAL 06'
        
        

        
        ws.border = Border(left=Side(border_style='thin'))
        ws.print_area = 'A1:I104'
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToHeight = 2
        ws.page_setup.fitToWidth = True
        
        
        wb.save(output_path30) 

    def toexcel_30comp(self):
        path_to_use30c = self.myapp.filename
        df_1 = pd.read_excel(path_to_use30c, 
            usecols=[2, 3, 5, 8, 16], 
            names=['cod', 'prod', 'qtd', 'cst_un', 'vnd_un'],
            skiprows=7)
        
        cod = df_1['cod']
        prod = df_1['prod']
        qtd = df_1['qtd']
        cst = df_1['cst_un']
        vnd = df_1['vnd_un']

        df_dict = {'cod' : cod, 
            'prod': prod,
            'qtd' : qtd,
            'cst_un': cst,
            'vnd' : vnd,
            'margem' : round(((vnd - cst)/vnd) * 100, 2), 
            'vnd30' : cst+((cst*30)/100),
            'vnd40': cst+((cst*40)/100)}

        df_30 = pd.DataFrame(df_dict)
        
        df_30['margem'] = df_30['margem'].astype(float)
        abaixo30 = df_30.sort_values(by='margem', ascending = True)
        abaixos = abaixo30[abaixo30['margem'] < 30] 
        output_path30 = filedialog.asksaveasfilename(initialdir="Desktop", title='Select Path to Save', filetypes = (('excel 2010', '*.xlsx'), ('all files', '*.*')), defaultextension='.xlsx')
        abaixos.to_excel(str(output_path30), header = ['Código', 'Produto', 'QTD', 'Custo', 'Venda', 'Margem', 'Venda 30%', 'Venda 40%'], startrow=2 )
        

        

        #styling
        wb = ol.load_workbook(output_path30)
        ws = wb.active
        
        


        relatorio = NamedStyle(name='relatorio')
        relatorio.font = Font(bold=True, name='Tahoma', size=16)
        bd = Side(style='thin', color="000000")
        relatorio.border = Border(left=bd, right=bd, top=bd, bottom=bd)
        top_title = ws['A1']
        top_title.fill
        



        for cell in ws['B']:
            cell.style = relatorio
            cell.number_format = 'General'
        for cell in ws['C']:
            cell.style = relatorio
        for cell in ws['D']:
            cell.style = relatorio
        for cell in ws['E']:
            cell.style = relatorio
            cell.number_format = '#,##0.00'
        for cell in ws['F']:
            cell.style = relatorio
            cell.number_format = '#,##0.00'
        for cell in ws['G']:
            cell.style = relatorio
            cell.number_format = '#,##0.00'
        for cell in ws['H']:
            cell.style = relatorio
            cell.number_format = '#,##0.00'
        for cell in ws['I']:
            cell.style = relatorio
            cell.number_format = '#,##0.00'
        

        ws.merge_cells('A1:I1')
        a1 = ws['A1']
        a1.font = Font(name='Arial', size=26, bold=True)
        c1 = ws['C1']
        ws.column_dimensions['B'].width = 13
        ws.column_dimensions['C'].width = 70
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 13
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 17
        ws.column_dimensions['I'].width = 23

        ws.merge_cells('A2:B2')
        ws.merge_cells('C2:I2')

        a2 = ws['A2']
        ws['A2'] = 'A/C: Diretor'
        a2.font = Font(name='Arial', size=15, bold=True)
        ws['C2'] = dt.datetime.today()
        c2=ws['C2']
        c2.font = Font(name='Arial', size=15, bold=True)
        c2.number_format = 'dd-mm-yyyy'
        ws['A1'] = '→30% Comparado  -  SOBERANO FILIAL 06'
        
        

        
        ws.border = Border(left=Side(border_style='thin'))
        ws.print_area = 'A1:I104'
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToHeight = 2
        ws.page_setup.fitToWidth = True
        
        
        wb.save(output_path30) 




root = Tk()
my_gui = Application(root)

root.mainloop()
